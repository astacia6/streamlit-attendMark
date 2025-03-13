import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import zipfile
import os
import time
import re

# 페이지 설정
st.set_page_config(page_title="출결 서류 처리", layout="wide")

# 제목 및 설명 추가
st.title("출결 서류 항목 표시")
st.markdown("""
    <style>
    .main {background-color: #f0f2f6;}
    </style>
    """, unsafe_allow_html=True)

st.markdown("### 반별 출결 엑셀 파일을 업로드하세요.")
st.markdown("업로드된 파일에서 '출석인정' 또는 '결석'이 포함된 셀을 음영 처리합니다.")

# 레이아웃 조정
col1, col2 = st.columns([1, 3])

with col1:
    # 파일 업로드
    uploaded_files = st.file_uploader("엑셀 파일 선택", type=["xlsx"], accept_multiple_files=True)

if uploaded_files:
    # 결과 파일을 저장할 임시 디렉토리 생성
    os.makedirs("temp", exist_ok=True)
    
    # 파일 이름에서 월 정보를 추출
    month_match = None
    for uploaded_file in uploaded_files:
        match = re.search(r'(\d{2})월', uploaded_file.name)
        if match:
            month_match = match.group(1)
            break

    for uploaded_file in uploaded_files:
        # 엑셀 파일 읽기
        workbook = load_workbook(uploaded_file)
        sheet = workbook.active

        # 특정 셀의 값을 검사하여 조건에 맞는 셀에 음영 처리
        fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            cell = row[4]  # E열 (0부터 시작하므로 인덱스 4)
            cell_value = str(cell.value)
            if ("출석인정" in cell_value or "결석" in cell_value) and "미인정결석" not in cell_value:
                cell.fill = fill

        # 병합된 셀에서 B6의 값을 가져오기
        b6_value = sheet['B6'].value
        n6_value = sheet['N6'].value

        # n6_value를 문자열로 변환하여 마지막 두 글자 추출
        if b6_value and n6_value:
            n6_str = str(n6_value)
            timestamp = int(time.time())
            new_filename = f"temp/{b6_value}{n6_str[-2:]}월_{timestamp}.xlsx"
        else:
            new_filename = f"temp/highlighted_{uploaded_file.name}_{timestamp}.xlsx"

        # 엑셀 파일을 다시 저장
        workbook.save(new_filename)

    # 추출한 월 정보를 사용하여 ZIP 파일 이름 설정
    if month_match:
        zip_filename = f"{month_match}월별출결현황.zip"
    else:
        zip_filename = "출결현황.zip"

    # ZIP 파일로 압축
    with zipfile.ZipFile(zip_filename, 'w') as zipf:
        for file in os.listdir("temp"):
            zipf.write(os.path.join("temp", file), file)

    # 안내 문구 및 이모지 추가
    st.markdown("<h3>출결 필수 서류 항목 표시가 완료되었습니다.</h3>", unsafe_allow_html=True)

    # 다운로드 링크 제공
    with open(zip_filename, "rb") as file:
        btn = st.download_button(
            label="다운로드",
            data=file,
            file_name=zip_filename,
            mime="application/zip"
        )

    # 임시 디렉토리 정리
    for file in os.listdir("temp"):
        os.remove(os.path.join("temp", file))
    os.rmdir("temp") 
